import io
import openpyxl
import logging
from datetime import datetime
from excel_sync.models import ExcelSyncSettings

logger = logging.getLogger(__name__)

class ExcelOnlineAdapter:
    """
    Data Source Adapter converting Excel Online .xlsx binary content into
    normalized record dicts for OrderImportService while preserving raw row payload and row numbers.
    """

    def __init__(self, settings_record: ExcelSyncSettings):
        self.settings = settings_record

    def parse_workbook(self, xlsx_bytes: bytes) -> list[dict]:
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        
        if self.settings.sheet_name and self.settings.sheet_name in workbook.sheetnames:
            sheet = workbook[self.settings.sheet_name]
        else:
            sheet = workbook.active

        header_row_idx = max(1, self.settings.header_row)
        rows = list(sheet.iter_rows(values_only=True))

        if not rows or len(rows) < header_row_idx:
            logger.warning("Excel workbook contains no rows or header row is beyond data length.")
            return []

        raw_headers = rows[header_row_idx - 1]
        headers = [str(h).strip() if h is not None else f"Column_{i+1}" for i, h in enumerate(raw_headers)]

        column_mapping = self.settings.column_mapping or {}
        normalized_records = []
        data_rows = rows[header_row_idx:]

        for row_offset, row in enumerate(data_rows, start=header_row_idx + 1):
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue

            record = {}
            extra_attributes = {}
            raw_row_dict = {}

            for idx, cell_val in enumerate(row):
                if idx >= len(headers):
                    continue

                header_name = headers[idx]
                target_field = column_mapping.get(header_name)
                cleaned_val = self._clean_cell_value(cell_val)
                raw_row_dict[header_name] = cleaned_val

                if target_field:
                    record[target_field] = cleaned_val
                else:
                    if header_name and not header_name.startswith("Column_"):
                        extra_attributes[header_name] = cleaned_val

            if extra_attributes:
                record['extra_attributes'] = extra_attributes

            record['_row_number'] = row_offset
            record['_raw_payload'] = raw_row_dict
            normalized_records.append(record)

        return normalized_records

    @staticmethod
    def inspect_headers(xlsx_bytes: bytes, header_row: int = 1) -> list[str]:
        workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows or len(rows) < header_row:
            return []

        header_vals = rows[header_row - 1]
        return [str(h).strip() for h in header_vals if h is not None and str(h).strip() != '']

    def _clean_cell_value(self, cell_val):
        if cell_val is None:
            return None
        if isinstance(cell_val, datetime):
            return cell_val.isoformat()
        if isinstance(cell_val, (int, float)):
            if isinstance(cell_val, float) and cell_val.is_integer():
                return str(int(cell_val))
            return cell_val
        return str(cell_val).strip()
